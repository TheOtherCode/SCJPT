# -*- coding: utf-8 -*-
"""
Created on Mon Jan 22 07:37:34 2024

@author: Vedant
"""


'''
This Block of Code access the Master List and creates an excel sheet within the workbook for each judge mentioned in the Master List

Run the code only if such an excel sheet does not exist.

Excel has limitations on the number of Characters to be used for naming a sheet.
The compiler may raise an error - But the code should continue to run.
'''

import openpyxl

def create_sheets_from_master_list():
    """Opens an Excel workbook named 'Judges.xlsx', reads names from a sheet named 'Judge_MasterList',
       and creates new sheets with those names in the same workbook.
    """

    workbook = openpyxl.load_workbook("Judges.xlsx")
    master_worksheet = workbook["Judge_MasterList"]

    names = []
    for row in master_worksheet.iter_rows(min_row=2):  # Start from row 2 to skip headers
        name = row[0].value
        if name:  # Check for empty cells
            names.append(name)

    # Create new sheets with unique names
    for name in set(names):  # Use a set to avoid duplicates
        if name not in workbook.sheetnames:  # Check if sheet already exists
            workbook.create_sheet(name)

    workbook.save("Judges.xlsx")
    print("New sheets created successfully!")

create_sheets_from_master_list()

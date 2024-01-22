# -*- coding: utf-8 -*-
"""
Created on Mon Jan 22 07:32:20 2024

@author: Vedant
"""

'''
This Block of Code will:
1. Access Judges Directory
2. Create a MaterList in a previously created Excel Sheet


Req:-
1. Create an empty Excel Sheet Book before running the code
'''

'''DO NOT RUN FOR EVERY JUDGE - ONLY TO BE RUN IF MASTER LIST TO BE UPDATED OR REFORMED'''



from selenium import webdriver
from selenium.webdriver.common.keys import Keys
driver = webdriver.Chrome()
def judge_dates():
    
    driver.get('https://main.sci.gov.in/chief-justice-judges')
    temp = (driver.find_element("id", 'CJ')).text
    judge_and_term = temp.split('\n')
           
    
    judge_and_term.remove('Hon\'ble Dr. Justice D.Y. Chandrachud')
    judge_and_term.remove('Hon\'ble Judges')
    judge_and_term.remove('The Chief Justice of India')
    judge_and_term.remove('DoB - Date of Birth')
    judge_and_term.remove('DoA - Date of Appointment')
    judge_and_term.remove('DoR - Date of Retirement')
    judge_and_term.remove('(DoB.) 11-11-1959')
    DoR = []
    DoA = []
    judges_final = ['Hon\'ble the Chief Justice']

    for i in judge_and_term:
        
        if i == 'PROFILE':
            continue
        if '(DoB.)' in i:
            judges_final.append(i[:-18])
        if 'Term' in i:
            judge_and_term.remove(i)
            DoA.append(i[22:32])
            DoR.append(i[42:])
         
    raw = [judges_final, DoA, DoR]
    return raw
    

import openpyxl


'''
The following method enters the data recieved from the above method (judges_dates) into an Excel Sheet.
It prepares a MasterList of All sitting Judges - 
Name of Judges
Date of Appointment
Date of Retirement
'''

def fill_excel_sheet(list1, list2, list3):
    
    '''
    Args:
        list1 (list): Takes values for names of judges.
        list2 (list): Takes values for the dates of appointment.
        list3 (list): Takes values for the dates of retirement.
    ''' 

    workbook = openpyxl.load_workbook("Judges.xlsx")
    worksheet = workbook["Judge_MasterList"]

    for i in range(len(list1)):
        worksheet.cell(row=i+2, column=1).value = list1[i]
        worksheet.cell(row=i+2, column=2).value = list2[i]
        worksheet.cell(row=i+2, column=3).value = list3[i]

    workbook.save("Judges.xlsx")

# Example usage:

data = judge_dates()

fill_excel_sheet(data[0], data[1], data[2])

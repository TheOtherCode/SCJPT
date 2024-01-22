# -*- coding: utf-8 -*-
"""
Created on Mon Jan 22 07:41:41 2024

@author: Vedant
"""

'''
This Block of Code will access all orders - only passed by a judge

Requirements:-
1. update judge xpath/id of Specific judge [Option Number]
2. May not work on the first try [May give some sort of an error - only because loading the html on the page takes a time]
3. Should work smoothly from the second try onwards.
4. Code has been written for Google Chrome Driver [Change the relevant lines, if you are using a different browser]
5. A mandatory timer of 120 seconds has been added to wait for the page html to load.
6. User will have to run this code for every single judge seperately

7. THIS IS GOING TO BE SLOW!

8. Info to be inserted in code - 
    a. line number 41 - the id tag for the judge selected
    b. line number 46 - date of appointment of the judge
     
'''


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()

driver.get('https://main.sci.gov.in/daily-order')
#Locating and Clicking on Search by Judge Name subtab
judge_tab = driver.find_element('xpath', '//*[@id="tabbed-nav"]/ul[2]/li[4]/a')
judge_tab.click()
time.sleep(2)
judge = driver.find_element('xpath', 'insert ID tag of judge here')

judge.click()
date1 = driver.find_element('xpath', '//*[@id="from_date"]')
date1.clear()
date1.send_keys('date of app')
date2 = driver.find_element('xpath', '//*[@id="to_date"]')
#date2.send_keys('16-01-2024')
# Locating Captcha value and removing its WhiteSpaces
captcha_text = driver.find_element('id', 'cap').text.strip()
#Locating Captcha enter_text box
captcha_box = driver.find_element('id', 'ansCaptcha')
captcha_box.send_keys(captcha_text)
#Locating and Clicking on Captcha submit button
submit = driver.find_element('id', 'getJBD')
submit.click()

#element = WebDriverWait(driver, 10).until(EC.presence_of_element_located(('xpath', '//*[@id="JBD"]/table/tbody/tr[11]/td[3]/a')))
time.sleep(120)
judgments = driver.find_elements('xpath', '//*[@id="JBD"]/table/tbody')[0].text


'''
The following block of code should print out a certain value.
Cross check if the value is the same as the last serial number on the page open on the corresponding google chrome window
'''
j = judgments.split('\n')
data = []

for i in j:
    if 'Case Number' in i:
        data.append(i)
#print(data)
thedate = []
for i in data:
    thedate.append(i[-10:])
    
print(len(thedate))
from collections import Counter
count = Counter(thedate)


'''
The follwing block of code takes the list from previously and pushes it onto excel sheet

Make sure the name you want to put matches the name in the excel sheet

Make sure to change - 
    a. line number 99 - Copy paste name of the sheet where you want the data to land. Make sure it is the correct judge ('name')
'''


import openpyxl

def write(dict_data):
    

    workbook = openpyxl.load_workbook("Judges.xlsx")
    worksheet = workbook['name of sheet in the excel workbook i.e., sheet holding the name of judge']  # Get the specified worksheet

    # Write the keys and values to the sheet
    row_num = 2  # Start writing from row 2
    for key, value in dict_data.items():
        worksheet.cell(row=row_num, column=1).value = key
        worksheet.cell(row=row_num, column=2).value = value
        row_num += 1

    workbook.save('Judges.xlsx')  # Save the workbook

write(count)
